import openpyxl as xl
import re
from openpyxl.chart import BarChart, Reference

def process_excel_file(filename):
    # Load the workbook
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # Loop through column C and print its values
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        print(cell.value)

    # Loop through column C, adjust values, and store them in column D
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        # Remove non-numeric characters (including currency symbol) using regex
        value_str = re.sub(r'[^\d.]', '', cell.value)
        # Convert cleaned value to float and then multiply by 0.9
        correct_price = float(value_str) * 0.9
        correct_price_cell = sheet.cell(row, 4)
        correct_price_cell.value = correct_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row,
                      min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'F2')

    # Save the modified workbook
    wb.save(filename)

    # Close the workbook
    wb.close()

# Call the function with the filename
process_excel_file('Book1.xlsx')
